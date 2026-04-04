# ==========================================================
# [config.py]
# ⚠️ 이 주석 및 파일명 표기는 절대 지우지 마세요.
# ==========================================================
import json
import os
import datetime
import pytz
import logging
import math
import time
import shutil
import tempfile
import pandas_market_calendars as mcal
import queue
import threading

# 🔎 [V30-Unified] 알림용 비동기 로직 제거 (네이티브 가동으로 동기식 전환)
# notification_queue = queue.Queue() - REMOVED

# 🔎 [V25-Diagnostic] 설정 파일 로딩 경로 추적
logging.warning(f"🔎 [CONFIG-TRACE] ConfigManager 모듈 로드됨: {os.path.abspath(__file__)}")

try:
    from version_history import VERSION_HISTORY
except ImportError:
    VERSION_HISTORY = ["V14.x [-] 버전 기록 파일(version_history.py)을 찾을 수 없습니다."]

class ConfigManager:
    def __init__(self, is_real=None):
        # ⚠️ [MODE] 파일은 루트의 data/ 폴더에 고정 (UI의 기본 뷰 모드 설정 역할)
        self.MODE_FILE = "data/trading_mode.dat"

        # 🌐 [V23.1 Dual-Core] 인스턴스 생성 시 모드를 강제 지정하거나 파일에서 읽음
        if is_real is not None:
            self.is_real = bool(is_real)
        else:
            self.is_real = self._load_mode_file()
            
        self._ensure_dirs()
        
        self.FILES = {
            "TOKEN": "token.dat",
            "CHAT_ID": "chat_id.dat",
            "LEDGER": "manual_ledger.json",    
            "HISTORY": "manual_history.json",  
            "SPLIT": "split_config.json",
            "TICKER": "active_tickers.json",
            "TURBO": "turbo_mode.dat",
            "ENGINE_STATUS": "engine_status.dat", # 🚀 [V22.3] 엔진 가동 상태 (ON/OFF)
            "SECRET_MODE": "secret_mode.dat",
            "PROFIT_CFG": "profit_config.json",
            "LOCKS": "trade_locks.json",
            "SEED_CFG": "seed_config.json",         
            "COMPOUND_CFG": "compound_config.json",
            "VERSION_CFG": "version_config.json",
            "REVERSE_CFG": "reverse_config.json",
            "ACTIVE_SEED_CFG": "active_seed.json",  # 🌐 [V23.1 Dual-Core] 현재 사이클용 액티브 시드 (qty > 0일 때 고정)
            "SNIPER_MULTIPLIER_CFG": "sniper_multiplier.json",
            "SPLIT_HISTORY": "split_history.json",
            "PORTFOLIO_RATIO": "portfolio_ratio.json",
            "SNAPSHOTS": "daily_snapshots.json",
            "CAPITAL": "capital_flow.json",
            "LIVE_STATUS": "live_status.json",
            "GLOBAL_TACTICS": "global_tactics.json", # 🛡️ [V25] 글로벌 전술 설정
            "SHADOW_BOUNCE": "shadow_bounce.dat",   # 🌓 [V25] 전술 상세 속성 분리
            "ENGINE_STATUS": "engine_status.dat",
            "REFRESH_TRIGGER": "refresh_needed.tmp",
            "EVENT_LOG": "event_log.json",
            "NOTIFICATIONS": "notifications.json"
        }
        
        self.DEFAULT_SEED = {"SOXL": 6720.0, "TQQQ": 6720.0}
        self.DEFAULT_SPLIT = {"SOXL": 40.0, "TQQQ": 40.0}
        self.DEFAULT_TARGET = {"SOXL": 12.0, "TQQQ": 10.0}
        self.DEFAULT_COMPOUND = {"SOXL": 70.0, "TQQQ": 70.0}
        self.DEFAULT_VERSION = {"SOXL": "V14", "TQQQ": "V14"}
        self.DEFAULT_PORTFOLIO_RATIO = {"SOXL": 0.55, "TQQQ": 0.45} # 🌟 [V22 패치] 100% 동적분할 기본 타겟 비중 (예비비 0%)
        
        self.DEFAULT_SNIPER_MULTIPLIER = {"SOXL": 1.0, "TQQQ": 0.9}

    def _get_base_dir(self):
        """🌟 [V33.2] 절대 경로 강제 (WSL 환경 데이터 격리 보장)"""
        if os.name == 'nt':
            root = "data"
        else:
            root = "/home/jmyoon312/data"
        
        base = os.path.join(root, "real" if self.is_real else "mock")
        if not os.path.exists(base):
            try: os.makedirs(base, exist_ok=True)
            except: pass
        return base

    def _ensure_dirs(self):
        """기본 디렉토리 생성 보장"""
        root = "data"
        if not os.path.exists(root): os.makedirs(root, exist_ok=True)
        self._get_base_dir()

    def _get_file_path(self, key):
        """파일명 키값에 대해 현재 모드의 절대 경로를 반환합니다."""
        base = self._get_base_dir()
        root = "data"

        if key in self.FILES:
            return os.path.join(base, self.FILES[key])
        
        # 🧪 [V33.1] 하드코딩 레거시 대응 지점
        if key == "MOCK_LOC_STAGING":
            return os.path.join(base, "mock_loc_staged.json")
            
        # 그 외는 루트 data/ 폴더 (공용 설정 등)
        return os.path.join(root, key)

    def _load_json(self, key, default=None):
        filename = self._get_file_path(key)
        if os.path.exists(filename):
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if data is None: 
                        return default if default is not None else {}
                    # 🛡️ [V23.5.1] 타입 안정성 강화: 로드된 데이터와 기본값의 타입이 다르면 기본값 반환
                    if default is not None and type(data) != type(default):
                        return default
                    return data
            except Exception as e:
                print(f"⚠️ [Config] JSON 로드 에러 ({filename}): {e}")
                try:
                    import shutil
                    shutil.copy(filename, filename + f".bak_{int(time.time())}")
                except Exception as backup_e:
                    print(f"⚠️ [Config] 백업 실패: {backup_e}")
                return default if default is not None else {}
        return default if default is not None else {}


    def get_ratio(self, ticker):
        """💰 [V24] 종목별 포트폴리오 비중 획득 (기본값: 균등 배분)"""
        try:
            ratios = self._load_json("PORTFOLIO_RATIO", {})
            if not ratios:
                return self.DEFAULT_PORTFOLIO_RATIO.get(ticker, 0.5)
            
            val = ratios.get(ticker)
            if val is None:
                return self.DEFAULT_PORTFOLIO_RATIO.get(ticker, 0.5)
            return float(val)
        except Exception:
            return 0.5

    def _save_json(self, key, data):
        filename = self._get_file_path(key)
        try:
            dir_name = os.path.dirname(filename)
            if dir_name and not os.path.exists(dir_name):
                os.makedirs(dir_name, exist_ok=True)
                
            fd, temp_path = tempfile.mkstemp(dir=dir_name, text=True)
            try:
                with os.fdopen(fd, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                    f.flush()
                    os.fsync(f.fileno())
                os.replace(temp_path, filename)
            except Exception as e:
                if os.path.exists(temp_path): os.remove(temp_path)
                raise e
        except Exception as e:
            print(f"❌ [Config] JSON 저장 중 치명적 에러 발생 ({filename}): {e}")
            if 'temp_path' in locals() and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass

    def _load_file(self, key, default=None):
        filename = self._get_file_path(key)
        if os.path.exists(filename):
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    return f.read().strip()
            except Exception as e:
                print(f"⚠️ [Config] 파일 로드 에러 ({filename}): {e}")
        return default

    def _save_file(self, key, content):
        filename = self._get_file_path(key)
        try:
            dir_name = os.path.dirname(filename)
            if dir_name and not os.path.exists(dir_name):
                os.makedirs(dir_name, exist_ok=True)
                
            fd, temp_path = tempfile.mkstemp(dir=dir_name, text=True)
            with os.fdopen(fd, 'w', encoding='utf-8') as f:
                f.write(str(content))
                f.flush()
                os.fsync(fd)
            os.replace(temp_path, filename)
        except Exception as e:
            print(f"❌ [Config] 텍스트 파일 저장 에러 ({filename}): {e}")

    def get_last_split_date(self, ticker):
        return self._load_json("SPLIT_HISTORY", {}).get(ticker, "")

    def set_last_split_date(self, ticker, date_str):
        d = self._load_json("SPLIT_HISTORY", {})
        d[ticker] = date_str
        self._save_json("SPLIT_HISTORY", d)

    def get_ledger(self):
        return self._load_json("LEDGER", [])

    def get_escrow_cash(self, ticker):
        locks = self._load_json("LOCKS", {})
        return float(locks.get(f"ESCROW_{ticker}", 0.0))

    def set_escrow_cash(self, ticker, amount):
        locks = self._load_json("LOCKS", {})
        locks[f"ESCROW_{ticker}"] = float(amount)
        self._save_json("LOCKS", locks)

    def add_escrow_cash(self, ticker, amount):
        current = self.get_escrow_cash(ticker)
        self.set_escrow_cash(ticker, current + float(amount))

    def clear_escrow_cash(self, ticker):
        locks = self._load_json("LOCKS", {})
        if f"ESCROW_{ticker}" in locks:
            del locks[f"ESCROW_{ticker}"]
            self._save_json("LOCKS", locks)

    def get_total_locked_cash(self, exclude_ticker=None):
        locks = self._load_json("LOCKS", {})
        total = 0.0
        for k, v in locks.items():
            if k.startswith("ESCROW_"):
                ticker_in_lock = k.replace("ESCROW_", "")
                if ticker_in_lock != exclude_ticker:
                    total += float(v)
        return total

    def is_locked(self, ticker, market_type):
        locks = self._load_json("TRADE_LOCKS", {})
        key = f"{ticker}_{market_type}"
        return locks.get(key, False)

    def set_lock(self, ticker, market_type):
        locks = self._load_json("TRADE_LOCKS", {})
        key = f"{ticker}_{market_type}"
        locks[key] = True
        self._save_json("TRADE_LOCKS", locks)

    def clear_lock(self, ticker):
        locks = self._load_json("TRADE_LOCKS", {})
        keys_to_del = [k for k in locks if k.startswith(f"{ticker}_")]
        for k in keys_to_del:
            del locks[k]
        self._save_json("TRADE_LOCKS", locks)

    def get_absolute_t_val(self, ticker, actual_qty, actual_avg_price):
        seed = float(self.get_seed(ticker))
        # 🎯 [V24] 시드 0 방어 로직 (시드 설정 누락 시 기본값 1000달러 기준 보정)
        if seed <= 0: seed = 1000.0
        
        split = float(self.get_split_count(ticker))
        if split <= 0: split = 40.0
        
        one_portion = seed / split
        # 🎯 [V24] T값 이상 폭주 방어 (정상적인 1분량 기준 t_val 계산)
        t_val = (actual_qty * actual_avg_price) / one_portion if one_portion > 0 else 0.0
        return round(t_val, 4), one_portion

    def apply_stock_split(self, ticker, ratio):
        if ratio <= 0: return
        ledger = self.get_ledger()
        changed = False
        for r in ledger:
            if r.get('ticker') == ticker:
                new_qty = round(r['qty'] * ratio)
                r['qty'] = new_qty if new_qty > 0 else (1 if r['qty'] > 0 else 0)
                r['price'] = round(r['price'] / ratio, 4)
                if 'avg_price' in r:
                    r['avg_price'] = round(r['avg_price'] / ratio, 4)
                changed = True
        if changed:
            self._save_json("LEDGER", ledger)

    def overwrite_genesis_ledger(self, ticker, genesis_records, actual_avg):
        ledger = self.get_ledger()
        remaining = [r for r in ledger if r['ticker'] != ticker]
        
        for i, rec in enumerate(genesis_records):
            remaining.append({
                "id": i + 1,
                "date": rec['date'],
                "ticker": ticker,
                "side": rec['side'],
                "price": rec['price'],
                "qty": rec['qty'],
                "avg_price": actual_avg, 
                "exec_id": f"GENESIS_{int(time.time())}_{i}",
                "is_reverse": False 
            })
        self._save_json("LEDGER", remaining)

    def overwrite_incremental_ledger(self, ticker, temp_recs, new_today_records):
        ledger = self.get_ledger()
        remaining = [r for r in ledger if r['ticker'] != ticker]
        updated_ticker_recs = list(temp_recs)
        
        current_rev_state = self.get_reverse_state(ticker).get("is_active", False)
        max_id = max([r.get('id', 0) for r in ledger] + [0])
        
        for i, rec in enumerate(new_today_records):
            max_id += 1
            new_row = {
                "id": max_id,
                "date": rec['date'],
                "ticker": ticker,
                "side": rec['side'],
                "price": rec['price'],
                "qty": rec['qty'],
                "avg_price": rec['avg_price'],
                "exec_id": rec.get("exec_id", f"FASTTRACK_{int(time.time())}_{i}"),
                "is_reverse": current_rev_state
            }
            if "desc" in rec:
                new_row["desc"] = rec["desc"]
                
            updated_ticker_recs.append(new_row)
            
        remaining.extend(updated_ticker_recs)
        self._save_json("LEDGER", remaining)

    def overwrite_ledger(self, ticker, actual_qty, actual_avg):
        ledger = self.get_ledger()
        remaining = [r for r in ledger if r['ticker'] != ticker]
        
        kst = pytz.timezone('Asia/Seoul')
        today_str = datetime.datetime.now(kst).strftime('%Y-%m-%d')
        new_id = 1 if not remaining else max(r.get('id', 0) for r in remaining) + 1
        
        remaining.append({
            "id": new_id, "date": today_str, "ticker": ticker, "side": "BUY",
            "price": actual_avg, "qty": actual_qty, "avg_price": actual_avg, 
            "exec_id": f"INIT_{int(time.time())}", "desc": "초기동기화", "is_reverse": False
        })
        self._save_json("LEDGER", remaining)

    def calibrate_avg_price(self, ticker, actual_avg):
        ledger = self.get_ledger()
        target_recs = [r for r in ledger if r['ticker'] == ticker]
        if target_recs:
            for r in target_recs:
                r['avg_price'] = actual_avg
            self._save_json("LEDGER", ledger)

    def clear_ledger_for_ticker(self, ticker):
        ledger = self.get_ledger()
        remaining = [r for r in ledger if r['ticker'] != ticker]
        self._save_json("LEDGER", remaining)
        self.set_reverse_state(ticker, False, 0, 0.0)
        self.clear_escrow_cash(ticker)

    def calculate_holdings(self, ticker, records=None):
        if records is None:
            records = self.get_ledger()
        target_recs = [r for r in records if r['ticker'] == ticker]
        total_qty, total_invested, total_sold = 0, 0.0, 0.0    
        
        for r in target_recs:
            if r['side'] == 'BUY':
                total_qty += r['qty']
                total_invested += (r['price'] * r['qty'])
            elif r['side'] == 'SELL':
                total_qty -= r['qty']
                total_sold += (r['price'] * r['qty'])
        
        total_qty = max(0, int(total_qty))
        invested_up = math.ceil(total_invested * 100) / 100.0
        sold_up = math.ceil(total_sold * 100) / 100.0
        
        if total_qty == 0:
            avg_price = 0.0
        else:
            avg_price = 0.0
            if target_recs:
                avg_price = float(target_recs[-1].get('avg_price', 0.0))
                if avg_price == 0.0:
                    buy_sum = sum(r['price']*r['qty'] for r in target_recs if r['side']=='BUY')
                    buy_qty = sum(r['qty'] for r in target_recs if r['side']=='BUY')
                    if buy_qty > 0:
                        avg_price = buy_sum / buy_qty
        
        return total_qty, avg_price, invested_up, sold_up

    def get_reverse_state(self, ticker):
        d = self._load_json("REVERSE_CFG", {})
        return d.get(ticker, {"is_active": False, "day_count": 0, "exit_target": 0.0, "last_update_date": ""})

    def set_reverse_state(self, ticker, is_active, day_count, exit_target=0.0, last_update_date=None):
        if last_update_date is None:
            est = pytz.timezone('US/Eastern')
            last_update_date = datetime.datetime.now(est).strftime('%Y-%m-%d')
            
        d = self._load_json("REVERSE_CFG", {})
        d[ticker] = {"is_active": is_active, "day_count": day_count, "exit_target": exit_target, "last_update_date": last_update_date}
        self._save_json("REVERSE_CFG", d)

    def update_reverse_day_if_needed(self, ticker):
        return False

    def increment_reverse_day(self, ticker):
        state = self.get_reverse_state(ticker)
        if state.get("is_active"):
            est = pytz.timezone('US/Eastern')
            now_est = datetime.datetime.now(est)
            today_est_str = now_est.strftime('%Y-%m-%d')
            
            if state.get("last_update_date") != today_est_str:
                nyse = mcal.get_calendar('NYSE')
                is_trading_day = not nyse.schedule(start_date=now_est.date(), end_date=now_est.date()).empty
                
                if is_trading_day:
                    new_day = state.get("day_count", 0) + 1
                    self.set_reverse_state(ticker, True, new_day, state.get("exit_target", 0.0), today_est_str)
                    return True
                else:
                    self.set_reverse_state(ticker, True, state.get("day_count", 0), state.get("exit_target", 0.0), today_est_str)
                    return False
        return False

    def calculate_v14_state(self, ticker):
        ledger = self.get_ledger()
        target_recs = sorted([r for r in ledger if r['ticker'] == ticker], key=lambda x: x.get('id', 0))
        
        seed = self.get_seed(ticker)
        split = self.get_split_count(ticker)
        base_portion = seed / split if split > 0 else 1
        
        holdings = 0
        rem_cash = seed
        total_invested = 0.0
        
        for r in target_recs:
            if holdings == 0:
                rem_cash = seed
                total_invested = 0.0
                
            qty = r['qty']
            amt = qty * r['price']
            
            if r['side'] == 'BUY':
                rem_cash -= amt
                holdings += qty
                total_invested += amt
                
            elif r['side'] == 'SELL':
                if qty >= holdings: 
                    holdings = 0
                    rem_cash = seed
                    total_invested = 0.0
                else: 
                    if holdings > 0:
                        avg_price = total_invested / holdings
                        total_invested -= (qty * avg_price)
                    holdings -= qty
                    rem_cash += amt
                    
        avg_price = total_invested / holdings if holdings > 0 else 0.0
        # 🎯 [V24] V14 상태 계산 시에도 절대 T값 로직과 일치시킴
        t_val, _ = self.get_absolute_t_val(ticker, holdings, avg_price)
            
        if holdings > 0:
            safe_denom = max(1.0, split - t_val)
            current_budget = rem_cash / safe_denom
        else:
            current_budget = seed / split
            t_val = 0.0
            
        return max(0.0, round(t_val, 4)), max(0.0, current_budget), max(0.0, rem_cash)

    def archive_graduation(self, ticker, end_date, prev_close=0.0):
        ledger = self.get_ledger()
        target_recs = [r for r in ledger if r['ticker'] == ticker]
        if not target_recs:
            return None, 0
        
        ledger_qty, avg_price, _, _ = self.calculate_holdings(ticker, target_recs)
        
        if ledger_qty > 0:
            split = self.get_split_count(ticker)
            is_reverse = self.get_reverse_state(ticker).get("is_active", False)

            if is_reverse:
                divisor = 10 if split <= 20 else 20
                loc_qty = math.floor(ledger_qty / divisor)
            else:
                loc_qty = math.ceil(ledger_qty / 4)

            limit_qty = ledger_qty - loc_qty
            if limit_qty < 0: 
                loc_qty = ledger_qty
                limit_qty = 0

            target_ratio = self.get_target_profit(ticker) / 100.0
            target_price = math.ceil(avg_price * (1 + target_ratio) * 100) / 100.0
            loc_price = prev_close if prev_close > 0 else avg_price

            new_id = max((r.get('id', 0) for r in ledger), default=0) + 1

            if loc_qty > 0:
                rec_loc = {"id": new_id, "date": end_date, "ticker": ticker, "side": "SELL", "price": loc_price, "qty": loc_qty, "avg_price": avg_price, "exec_id": f"GRAD_LOC_{int(time.time())}", "is_reverse": is_reverse}
                ledger.append(rec_loc)
                target_recs.append(rec_loc)
                new_id += 1

            if limit_qty > 0:
                rec_limit = {"id": new_id, "date": end_date, "ticker": ticker, "side": "SELL", "price": target_price, "qty": limit_qty, "avg_price": avg_price, "exec_id": f"GRAD_LMT_{int(time.time())}", "is_reverse": is_reverse}
                ledger.append(rec_limit)
                target_recs.append(rec_limit)

            self._save_json("LEDGER", ledger)

        total_buy = math.ceil(sum(r['price']*r['qty'] for r in target_recs if r['side']=='BUY') * 100) / 100.0
        total_sell = math.ceil(sum(r['price']*r['qty'] for r in target_recs if r['side']=='SELL') * 100) / 100.0
        
        profit = math.ceil((total_sell - total_buy) * 100) / 100.0
        yield_pct = math.ceil(((profit / total_buy * 100) if total_buy > 0 else 0.0) * 100) / 100.0
        
        compound_rate = self.get_compound_rate(ticker) / 100.0
        added_seed = 0
        if profit > 0 and compound_rate > 0:
            added_seed = math.floor(profit * compound_rate)
            current_seed = self.get_seed(ticker)
            self.set_seed(ticker, current_seed + added_seed)

        history = self._load_json("HISTORY", [])
        new_hist = {
            "id": len(history) + 1, "ticker": ticker, "end_date": end_date,
            "profit": profit, "yield": yield_pct, "revenue": total_sell, "invested": total_buy, "trades": target_recs
        }
        history.append(new_hist)
        self._save_json("HISTORY", history)
        
        self.clear_ledger_for_ticker(ticker)
        
        return new_hist, added_seed

    def get_full_version_history(self):
        return VERSION_HISTORY

    def get_version_history(self):
        return VERSION_HISTORY

    def get_latest_version(self):
        history = self.get_version_history()
        if history and len(history) > 0:
            latest_entry = history[-1]
            if isinstance(latest_entry, dict):
                return latest_entry.get("version", "V14.x")
            elif isinstance(latest_entry, str):
                return latest_entry.split(' ')[0] 
        return "V14.x"

    def get_history(self):
        """졸업(수익확정) 기록을 불러옵니다. manual_history.json이 없으면 ledger.json에서 추출을 시도합니다."""
        history = self._load_json("HISTORY", [])
        if not history:
            # ledger.json에서 '졸업' 또는 큰 규모의 'SELL' 항목을 찾아 history 대용으로 사용
            ledger = self._load_json("LEDGER", [])
            for item in ledger:
                if "졸업" in item.get("desc", "") or "익절" in item.get("desc", ""):
                    history.append({
                        "ticker": item.get("ticker", "UNKNOWN"),
                        "profit": item.get("profit", 0.0),
                        "yield": item.get("yield", 0.0),
                        "end_date": item.get("date", "").split(" ")[0]
                    })
        return history


    def record_daily_snapshot(self, total_cash, holdings_value):
        """매일 장 마감 시 자산 상태를 스냅샷으로 저장합니다. (비중 표시용 개별 종목 평가금 포함)"""
        snapshots = self._load_json("SNAPSHOTS", [])
        est = pytz.timezone('US/Eastern')
        today = datetime.datetime.now(est).strftime('%Y-%m-%d')
        
        # [V24] 실시간 상태에서 종목별 비중 정보를 가져와 함께 저장합니다.
        live_status = self._load_json("LIVE_STATUS", {})
        ticker_eval = {}
        if "tickers" in live_status:
            for t, info in live_status["tickers"].items():
                eval_val = info.get("qty", 0) * info.get("current_price", 0)
                if eval_val > 0: ticker_eval[t] = round(eval_val, 2)
        
        # holdings_value가 0이면 live_status에서 합산 시도 (자산 표시 버그 수정)
        if holdings_value <= 0 and ticker_eval:
            holdings_value = sum(ticker_eval.values())

        new_snap = {
            "date": today,
            "cash": round(total_cash, 2),
            "holdings": round(holdings_value, 2),
            "total": round(total_cash + holdings_value, 2),
            "ticker_eval": ticker_eval
        }
        
        idx = next((i for i, s in enumerate(snapshots) if s['date'] == today), -1)
        if idx >= 0:
            snapshots[idx] = new_snap
        else:
            snapshots.append(new_snap)
            
        self._save_json("SNAPSHOTS", snapshots)
        return new_snap


    def add_capital_flow(self, amount, flow_type="DEPOSIT"):
        """자본금 입출금 기록을 저장합니다 (수익률 계산 보정용)."""
        flows = self._load_json("CAPITAL", [])
        est = pytz.timezone('US/Eastern')
        today = datetime.datetime.now(est).strftime('%Y-%m-%d %H:%M:%S')
        
        flows.append({
            "date": today,
            "amount": float(amount),
            "type": flow_type
        })
        self._save_json("CAPITAL", flows)

    def get_analytics_data(self):
        """프론트엔드 분석 대시보드용 통합 데이터를 반환합니다 (실시간 상태 포함)."""
        snapshots = self._load_json("SNAPSHOTS", [])
        history = self.get_history()
        live_status = self._load_json("LIVE_STATUS", {})
        
        # 실시간 현재 자산 정보 추출 (대시보드 상단 바용)
        # live_status.json 구조: {"cash": 123, "holdings_value": 456, "tickers": {...}}
        current_valuation = {
            "cash": live_status.get("cash", 0),
            "holdings": live_status.get("holdings_value", 0),
            "total": live_status.get("cash", 0) + live_status.get("holdings_value", 0),
            "ticker_eval": {}
        }
        
        if "tickers" in live_status:
            for t, info in live_status["tickers"].items():
                eval_val = info.get("qty", 0) * info.get("current_price", 0)
                if eval_val > 0:
                    current_valuation["ticker_eval"][t] = round(eval_val, 2)

        # 🛡️ [Bug Fix] holdings_value가 0이면 live_status에서 합산 시도 (자산 표시 버그 수정)
        if current_valuation.get("holdings", 0) <= 0 and current_valuation.get("ticker_eval"):
            h_sum = sum(current_valuation["ticker_eval"].values())
            if h_sum > 0:
                current_valuation["holdings"] = round(h_sum, 2)
                current_valuation["total"] = current_valuation["cash"] + current_valuation["holdings"]

        # 성과 지표 계산
        pos_profits = [h.get('profit', 0) for h in history if h.get('profit', 0) > 0]
        neg_profits = [h.get('profit', 0) for h in history if h.get('profit', 0) < 0]
        
        win_count = len(pos_profits)
        total_trades = len(history)
        win_rate = (win_count / total_trades * 100) if total_trades > 0 else 0
        
        total_p = sum(pos_profits)
        total_l = abs(sum(neg_profits))
        profit_factor = (total_p / total_l) if total_l > 0 else (total_p if total_p > 0 else 0)
        
        # [V23.5] Advanced Analytics Calculation
        holding_periods = []
        ticker_performance = {}

        for h in history:
            # 1. Holding Period Analysis
            start_date = h.get("start_date")
            end_date = h.get("end_date")
            if start_date and end_date:
                try:
                    s_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                    e_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d")
                    holding_periods.append(max(0, (e_dt - s_dt).days))
                except: pass
            
            # 2. Ticker-wise Performance
            ticker = h.get("ticker", "UNKNOWN")
            if ticker not in ticker_performance:
                ticker_performance[ticker] = {"profit": 0, "count": 0, "win": 0}
            
            p = h.get("profit", 0)
            ticker_performance[ticker]["profit"] += p
            ticker_performance[ticker]["count"] += 1
            if p > 0: ticker_performance[ticker]["win"] += 1

        avg_holding = round(sum(holding_periods) / len(holding_periods), 1) if holding_periods else 0

        # [V23.5] Manual Sync Tracking
        live_status = self._load_json("LIVE_STATUS", {})
        last_sync = live_status.get("last_manual_sync")
        if not last_sync:
            # Fallback if field is missing but we want to show something
            last_sync = {"timestamp": 0, "status": "IDLE", "msg": "대기 중"}

        return {
            "current_valuation": current_valuation,
            "snapshots": snapshots,
            "capital_flows": self._load_json("CAPITAL", []),
            "history": history,
            "metrics": {
                "win_rate": round(win_rate, 2),
                "profit_factor": round(profit_factor, 2),
                "total_trades": total_trades,
                "gross_profit": round(total_p, 2),
                "gross_loss": round(total_l, 2),
                "avg_holding_days": avg_holding,
                "ticker_performance": ticker_performance
            },
            "last_manual_sync": last_sync,
            "cycles": self.get_cycle_analytics(), # 사이클 분석 데이터 추가
            "periodical": self.get_periodical_analytics(),
            "tax": self.calculate_tax_estimation(),
            "events": self.get_recent_events(limit=1000) # [V26.5] 상세 기록 조회를 위해 1000개로 한도 대폭 확장
        }

    def clear_events(self):
        """🧹 [V26.5] 일일 사이클 시작 시 이전 운영 기록을 초기화합니다."""
        self._save_json("EVENT_LOG", [])
        self._last_logs = {} # 캐시 초기화

    def get_recent_events(self, limit=1000):
        """최근 실행된 작업 로그를 반환합니다."""
        events = self._load_json("EVENT_LOG", [])
        return events[-limit:] if events else []

    def get_ledger_explorer_data(self):
        """장부 탐색기를 위한 평탄화된 모든 거래 내역 반환 (활성 + 과거)"""
        active_ledger = self.get_ledger()
        history = self.get_history()
        
        all_records = []
        
        # 1. 활성 장부 데이터 (현재 진행 중인 매수 분할 내역)
        for t in active_ledger:
            all_records.append({
                "date": t.get("date"),
                "ticker": t.get("ticker"),
                "side": t.get("side", "BUY"),
                "price": t.get("price"),
                "qty": t.get("qty"),
                "status": "ACTIVE",
                "note": t.get("desc", "분할 매수 진행 중")
            })
        
        # 2. 과거 졸업 내역 데이터 (이미 청산 완료된 내역)
        for h in history:
            ticker = h.get("ticker")
            trades = h.get("trades", [])
            for t in trades:
                all_records.append({
                    "date": t.get("date"),
                    "ticker": ticker,
                    "side": t.get("side", "BUY"),
                    "price": t.get("price"),
                    "qty": t.get("qty"),
                    "status": "GRADUATED",
                    "note": t.get("desc", "정상 졸업 청산")
                })
                
        # 날짜 내림차순 정렬
        all_records.sort(key=lambda x: str(x.get("date", "")), reverse=True)
        return all_records

    def get_cycle_analytics(self):
        """거래 내역을 '사이클(매수 시작~졸업)' 단위로 묶어서 요약 데이터를 생성합니다."""
        history = self.get_history()
        active_ledger = self.get_ledger()
        
        cycles = []
        
        # 1. 졸업(완료) 사이클 처리
        # HISTORY의 각 항목은 이미 하나의 졸업된 사이클을 나타냅니다.
        for h in history:
            trades = h.get("trades", [])
            start_date = trades[0].get("date") if trades else h.get("end_date")
            
            cycles.append({
                "ticker": h.get("ticker"),
                "status": "GRADUATED",
                "start_date": start_date,
                "end_date": h.get("end_date"),
                "invested": h.get("invested", 0),
                "revenue": h.get("revenue", 0),
                "profit": h.get("profit", 0),
                "yield": h.get("yield", 0),
                "trade_count": len(trades),
                "trades": trades
            })
            
        # 2. 활성(진행 중) 사이클 처리
        # 종목별로 그룹화하여 현재 진행 상태를 요약합니다.
        active_map = {}
        for t in active_ledger:
            ticker = t.get("ticker")
            if ticker not in active_map:
                active_map[ticker] = []
            active_map[ticker].append(t)
            
        for ticker, trades in active_map.items():
            total_buy = sum(r['price'] * r['qty'] for r in trades if r['side'] == 'BUY')
            total_sell = sum(r['price'] * r['qty'] for r in trades if r['side'] == 'SELL')
            net_invested = total_buy - total_sell
            
            start_date = trades[0].get("date") if trades else "N/A"
            
            cycles.append({
                "ticker": ticker,
                "status": "ACTIVE",
                "start_date": start_date,
                "end_date": "-",
                "invested": round(net_invested, 2),
                "revenue": 0,
                "profit": 0,
                "yield": 0,
                "trade_count": len(trades),
                "trades": trades
            })
            
        # 최신 시작일 순으로 정렬
        cycles.sort(key=lambda x: str(x.get("start_date", "")), reverse=True)
        return cycles

    def get_periodical_analytics(self):
        """기록된 졸업 데이터를 기반으로 연/월별 수익 현황을 집계합니다."""
        history = self.get_history()
        periodical = {}
        for h in history:
            date_str = h.get('end_date', '')
            if not date_str: continue
            try:
                dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                year = str(dt.year)
                month = f"{dt.month:02d}"
                
                if year not in periodical: periodical[year] = {"profit": 0, "months": {}}
                if month not in periodical[year]["months"]: periodical[year]["months"][month] = 0
                
                profit = h.get('profit', 0.0)
                periodical[year]["profit"] += profit
                periodical[year]["months"][month] += profit
            except: continue
        return periodical

    def calculate_tax_estimation(self):
        """🚀 [V29.7] 미국 주식 양도소득세(국내 고정 22%) 추산 엔진"""
        history = self.get_history()
        this_year = str(datetime.datetime.now().year)
        
        # 올해 귀속분 수익 합산
        yearly_profit = sum(h.get('profit', 0) for h in history if str(h.get('end_date', '')).startswith(this_year))
        
        # 국내 기준 기본 공제액 (250만원 / 약 $1,900~$2,000)
        # 사용자 요청에 따라 비율은 22% 고정
        allowance = 2000.0
        taxable = max(0, yearly_profit - allowance)
        estimated_tax = taxable * 0.22
        
        return {
            "year": this_year,
            "total_profit": round(yearly_profit, 2),
            "allowance": allowance,
            "taxable_profit": round(taxable, 2),
            "estimated_tax": round(estimated_tax, 2),
            "tax_rate": 0.22
        }

    def get_ledger_stats(self):
        """📊 [V29.7] 장부 요약 통계 산출 (프론트엔드 위젯용)"""
        history = self.get_history()
        active_ledger = self.get_ledger()
        
        # 1. 실현 손익 (졸업 내역 기준)
        total_realized_profit = sum(h.get('profit', 0) for h in history)
        total_revenue = sum(h.get('revenue', 0) for h in history)
        total_invested = sum(h.get('invested', 0) for h in history)
        
        wins = [h for h in history if h.get('profit', 0) > 0]
        losses = [h for h in history if h.get('profit', 0) <= 0]
        
        win_rate = (len(wins) / len(history) * 100) if history else 0
        
        # 2. 미실현 손익 (활성 대시보드 참고)
        # live_status.json에서 현재가 기준 평가금액 합산
        live_status = self._load_json("LIVE_STATUS", {})
        unrealized_profit = 0
        active_value = 0
        
        if "tickers" in live_status:
            for t, info in live_status["tickers"].items():
                qty = info.get("qty", 0)
                avg = info.get("avg_price", 0)
                curr = info.get("current_price", 0)
                if qty > 0:
                    unrealized_profit += (curr - avg) * qty
                    active_value += curr * qty

        # 3. 세금 추산
        tax_info = self.calculate_tax_estimation()
        
        return {
            "total_realized_profit": round(total_realized_profit, 2),
            "total_revenue": round(total_revenue, 2),
            "total_invested": round(total_invested, 2),
            "win_rate": round(win_rate, 1),
            "unrealized_profit": round(unrealized_profit, 2),
            "active_value": round(active_value, 2),
            "total_cycles": len(history),
            "tax_liability": tax_info["estimated_tax"],
            "net_profit": round(total_realized_profit - tax_info["estimated_tax"], 2)
        }

    def export_ledger_excel(self, output_path):
        """📁 [V29.7] 전문가급 엑셀 리포트 생성 (Pandas + openpyxl)"""
        import pandas as pd
        
        history = self.get_history()
        stats = self.get_ledger_stats()
        
        # 1. 요약 시트 데이터 준비
        summary_data = {
            "항목": ["총 실현 손익 (누적)", "총 투자 원금", "총 회수 금액", "총 졸업 횟수", "승률 (%)", "예상 세금 (22%)", "세후 순이익"],
            "금액/수치": [
                f"${stats['total_realized_profit']:,.2f}",
                f"${stats['total_invested']:,.2f}",
                f"${stats['total_revenue']:,.2f}",
                f"{stats['total_cycles']}회",
                f"{stats['win_rate']}%",
                f"${stats['tax_liability']:,.2f}",
                f"${stats['net_profit']:,.2f}"
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        
        # 2. 상세 내역 시트 데이터 준비
        detailed_rows = []
        for h in history:
            detailed_rows.append({
                "졸업일": h.get("end_date"),
                "종목": h.get("ticker"),
                "투자원금": h.get("invested"),
                "회수금액": h.get("revenue"),
                "실현손익": h.get("profit"),
                "수익률(%)": h.get("yield")
            })
        df_details = pd.DataFrame(detailed_rows)
        
        # 3. 엑셀 파일 작성
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='투자 요약', index=False)
            df_details.to_excel(writer, sheet_name='상세 거래 내역', index=False)
            
            # 스타일링 (openpyxl)
            workbook = writer.book
            
            # 요약 시트 스타일
            ws_summary = writer.sheets['투자 요약']
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center")
            
            for cell in ws_summary["1:1"]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 상세 시트 스타일 (수익/손실 색상 구분)
            ws_details = writer.sheets['상세 거래 내역']
            from openpyxl.styles import Font, Color
            red_font = Font(color="FF0000") # 수익: 빨강
            blue_font = Font(color="0000FF") # 손실: 파랑

            for row in range(2, len(detailed_rows) + 2):
                profit_val = ws_details.cell(row=row, column=5).value
                if profit_val is not None:
                    if float(profit_val) > 0:
                        ws_details.cell(row=row, column=5).font = red_font
                        ws_details.cell(row=row, column=6).font = red_font
                    elif float(profit_val) < 0:
                        ws_details.cell(row=row, column=5).font = blue_font
                        ws_details.cell(row=row, column=6).font = blue_font

            # 열 너비 자동 조정
            for sheetname in ['투자 요약', '상세 거래 내역']:
                ws = writer.sheets[sheetname]
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

        return output_path



    def check_lock(self, ticker, market_type):
        est = pytz.timezone('US/Eastern')
        today = datetime.datetime.now(est).strftime('%Y-%m-%d')
        locks = self._load_json("LOCKS", {})
        return locks.get(f"{today}_{ticker}_{market_type}", False)

    def set_lock(self, ticker, market_type):
        est = pytz.timezone('US/Eastern')
        today = datetime.datetime.now(est).strftime('%Y-%m-%d')
        locks = self._load_json("LOCKS", {})
        locks[f"{today}_{ticker}_{market_type}"] = True
        self._save_json("LOCKS", locks)

    def reset_locks(self):
        locks = self._load_json("LOCKS", {})
        surviving_locks = {k: v for k, v in locks.items() if k.startswith("ESCROW_")}
        self._save_json("LOCKS", surviving_locks)
        
    def reset_lock_for_ticker(self, ticker):
        est = pytz.timezone('US/Eastern')
        today = datetime.datetime.now(est).strftime('%Y-%m-%d')
        locks = self._load_json("LOCKS", {})
        
        keys_to_delete = [k for k in locks.keys() if k.startswith(f"{today}_{ticker}")]
        if keys_to_delete:
            for k in keys_to_delete:
                del locks[k]
            self._save_json("LOCKS", locks)
    
    def get_seed(self, t):
        """🌟 [V23.1] 사용자가 설정한 '예약(Reserved)' 시드 반환"""
        return float(self._load_json("SEED_CFG", self.DEFAULT_SEED).get(t, 6720.0))

    def set_seed(self, t, v): 
        d = self._load_json("SEED_CFG", self.DEFAULT_SEED)
        d[t] = v
        self._save_json("SEED_CFG", d)

    def get_active_seed(self, t):
        """🚀 [V23.1] 현재 매매 사이클에서 실제로 사용 중인 시드 반환"""
        v = self._load_json("ACTIVE_SEED_CFG")
        # 액티브 시드가 없으면 일단 일반 시드(Reserved)를 가져옴
        return float(v.get(t, self.get_seed(t)))

    def set_active_seed(self, t, v):
        """🚀 [V23.1] 현재 매매 사이클의 시드 저장 (보통 qty=0일 때 호출)"""
        d = self._load_json("ACTIVE_SEED_CFG")
        d[t] = v
        self._save_json("ACTIVE_SEED_CFG", d)

    def get_compound_rate(self, t):
        return float(self._load_json("COMPOUND_CFG", self.DEFAULT_COMPOUND).get(t, 70.0))

    def set_compound_rate(self, t, v):
        d = self._load_json("COMPOUND_CFG", self.DEFAULT_COMPOUND)
        d[t] = v
        self._save_json("COMPOUND_CFG", d)

    def get_version(self, t):
        return self._load_json("VERSION_CFG", self.DEFAULT_VERSION).get(t, "V14")

    def set_version(self, t, v):
        d = self._load_json("VERSION_CFG", self.DEFAULT_VERSION)
        d[t] = v
        self._save_json("VERSION_CFG", d)
        
    def get_portfolio_ratio(self, t):
        # 🌟 [V22 패치] 각 개별 종목이 전체 계좌 총액 중에서 가져가야 할 목표 할당 퍼센티지 (0.0 ~ 1.0)
        return float(self._load_json("PORTFOLIO_RATIO", self.DEFAULT_PORTFOLIO_RATIO).get(t, self.DEFAULT_PORTFOLIO_RATIO.get(t, 0.45)))

    def set_portfolio_ratio(self, t, v):
        d = self._load_json("PORTFOLIO_RATIO", self.DEFAULT_PORTFOLIO_RATIO)
        d[t] = float(v)
        self._save_json("PORTFOLIO_RATIO", d)
        
    def rebalance_seed_on_graduation(self, ticker, total_asset):
        """
        🌟 [V22 패치] 비동기 독립 리밸런싱 엔진
        해당 종목이 "0주"로 익절(졸업) 했을 때, 커져버린 'Total Asset(예비현금+전체주식총 가치)'에 
        자신의 고유 배정 비율을 곱하여 새로운 등치(시드머니)를 영구 갱신시킴. (계단식 복리 메커니즘)
        """
        ratio = self.get_portfolio_ratio(ticker)
        new_seed = math.floor(total_asset * ratio)
        
        # 안전 장치: 시드가 너무 비정상적으로 작거나 0이 되는 것을 방어
        if new_seed < 100:
            print(f"⚠️ [Config] 너무 작은 리벨런싱 시드 감지 (${new_seed}). 리밸런싱 생략.")
            return self.get_seed(ticker)
            
        self.set_seed(ticker, new_seed)
        print(f"🔄 [Config] {ticker} 비동기 졸업 리밸런싱 완료! {ratio*100}% 비율 적용 → 새 시드: ${new_seed:,.0f}")
        return new_seed

    def clone_config_from_mode(self, ticker, source_is_real):
        """🌟 [V23.3] 전략 이식 엔진: 소스 모드의 설정을 현재 인스턴스 모드로 복제"""
        from config import ConfigManager as CM
        source_manager = CM(is_real=source_is_real)
        
        try:
            # 1. 시드 복제 (예약된 시드)
            self.set_seed(ticker, source_manager.get_seed(ticker))
            # 2. 분할 횟수 복제
            self.set_split_count(ticker, source_manager.get_split_count(ticker))
            # 3. 목표 수익률 복제
            self.set_target_profit(ticker, source_manager.get_target_profit(ticker))
            # 4. 복리율 복제
            self.set_compound_rate(ticker, source_manager.get_compound_rate(ticker))
            # 5. 버전 복제
            self.set_version(ticker, source_manager.get_version(ticker))
            # 6. 포트폴리오 비중 복제
            self.set_portfolio_ratio(ticker, source_manager.get_portfolio_ratio(ticker))
            
            print(f"🧬 [Config-Implant] {ticker} 전략 이식 완료 ({'REAL' if source_is_real else 'MOCK'} ➡️ {'REAL' if self.is_real else 'MOCK'})")
            return True
        except Exception as e:
            print(f"❌ [Config-Implant] {ticker} 이식 실패: {e}")
            return False

    def get_split_count(self, t):
        return self._load_json("SPLIT", self.DEFAULT_SPLIT).get(t, 40.0)

    def set_split_count(self, t, v):
        d = self._load_json("SPLIT", self.DEFAULT_SPLIT)
        d[t] = float(v)
        self._save_json("SPLIT", d)

    def get_target_profit(self, t):
        return self._load_json("PROFIT_CFG", self.DEFAULT_TARGET).get(t, 10.0)
        
    def set_target_profit(self, t, v):
        d = self._load_json("PROFIT_CFG", self.DEFAULT_TARGET)
        d[t] = float(v)
        self._save_json("PROFIT_CFG", d)
        
    def get_sniper_multiplier(self, t):
        default_val = self.DEFAULT_SNIPER_MULTIPLIER.get(t, 1.0)
        return float(self._load_json("SNIPER_MULTIPLIER_CFG", self.DEFAULT_SNIPER_MULTIPLIER).get(t, default_val))
        
    def set_sniper_multiplier(self, t, v):
        d = self._load_json("SNIPER_MULTIPLIER_CFG", self.DEFAULT_SNIPER_MULTIPLIER)
        d[t] = float(v)
        self._save_json("SNIPER_MULTIPLIER_CFG", d)

    def get_turbo_mode(self):
        val = self._load_file("TURBO")
        # 🚀 [V23.1] 부스터 모드는 기본적으로 "ON"으로 작동 (저장된 값이 없으면 True 반환)
        if val is None: return True
        return val == 'True'

    def set_turbo_mode(self, v):
        self._save_file("TURBO", str(v))

    def set_secret_mode(self, v):
        self._save_file("SECRET_MODE", str(v))

    # 🌓 [V23.1] 눌림목 추적 (Shadow-Strike) 연동
    def get_shadow_strike(self):
        val = self._load_file("SHADOW_STRIKE")
        return val != 'False' # 기본값 ON

    def set_shadow_strike(self, v):
        self._save_file("SHADOW_STRIKE", str(v))

    def get_shadow_bounce(self):
        val = self._load_file("SHADOW_BOUNCE")
        return float(val) if val else 1.5 # 기본값 1.5%

    def set_shadow_bounce(self, v):
        self._save_file("SHADOW_BOUNCE", str(v))

    # 🎯 [V23.1] 스나이퍼 방어 (Sniper-Defense) 연동
    def get_sniper_defense(self):
        val = self._load_file("SNIPER_DEFENSE")
        return val != 'False' # 기본값 ON
    def set_sniper_defense(self, v):
        self._save_file("SNIPER_DEFENSE", str(v))

    # 🏹 [V25] 글로벌 전술 관리 (Tactical Command Center 연동)
    def get_global_tactics(self):
        default = {
            "shield": False,
            "shadow": False,
            "turbo": False,
            "sniper": False, 
            "jupjup": False
        }
        return self._load_json("GLOBAL_TACTICS", default)

    def set_global_tactics(self, tactics):
        self._save_json("GLOBAL_TACTICS", tactics)

    def get_active_tickers(self):
        return self._load_json("TICKER", ["SOXL", "TQQQ"])

    def set_active_tickers(self, v):
        self._save_json("TICKER", v)

    def get_chat_id(self): 
        v = self._load_file("CHAT_ID")
        return int(v) if v else None

    def set_chat_id(self, v):
        self._save_file("CHAT_ID", v)

    def _load_mode_file(self):
        """🌟 [V22.2] 루트 data/ 폴더에서 현재 매매 모드를 읽어옵니다."""
        if os.path.exists(self.MODE_FILE):
            try:
                with open(self.MODE_FILE, 'r', encoding='utf-8') as f:
                    return f.read().strip() == 'True'
            except Exception:
                pass
        return False

    def get_is_real_trading(self):
        """전역 매매 모드 반환 (UI 표시용)"""
        return self._load_mode_file()

    def set_is_real_trading(self, v):
        """전역 매매 모드 설정 및 파일 저장 (UI 표시용)"""
        try:
            if not os.path.exists("data"): os.makedirs("data")
            with open(self.MODE_FILE, 'w', encoding='utf-8') as f:
                f.write(str(bool(v)))
        except Exception as e:
            print(f"❌ [Config] 모드 파일 저장 에러: {e}")

    def get_engine_status(self):
        """🚀 [V22.3] 해당 인스턴스의 엔진 가동 상태 반환 (기본값 ON)"""
        v = self._load_file("ENGINE_STATUS", "ON")
        return v == "ON"

    def set_engine_status(self, is_on: bool):
        """🚀 [V23.1] 엔진 작동 상태 저장"""
        self._save_file("ENGINE_STATUS", "ON" if is_on else "OFF")

    def is_market_open(self):
        """🕒 [V23.1] 현재 미국 시장 개장 상태 및 페이즈 판별 (Config 공유용)"""
        try:
            est = pytz.timezone('US/Eastern')
            today = datetime.datetime.now(est)
            # 1차: 명확한 주말은 무조건 휴장 처리 (0:월 ~ 6:일)
            if today.weekday() >= 5: 
                return "WEEKEND"
                
            nyse = mcal.get_calendar('NYSE')
            schedule = nyse.schedule(start_date=today.date(), end_date=today.date())
            
            if not schedule.empty:
                return "OPEN"
            else:
                # 달력 데이터가 아예 비어있으면 진짜 공휴일
                return "HOLIDAY"
        except Exception as e:
            # 패키지 구버전 등 달력 조회 에러 시, 평일이면 무조건 개장으로 강제 처리
            print(f"⚠️ [Config] 달력 라이브러리 에러 발생. 평일이므로 강제 개장 처리합니다: {e}")
            return "OPEN"

    def record_daily_snapshot(self, total_cash, holdings_value, ticker_state=None):
        """
        🚀 [V24 Stable] 주말/야간 데이터 보전용 통합 스냅샷
        1. 기존 분석용 데이터(현금, 주식합계) 저장
        2. [V24] 복구용 슬림 상태(수량, 평단, 종가, 전일종가) 병합
        """
        snapshots = self._load_json("SNAPSHOTS", [])
        est = pytz.timezone('US/Eastern')
        today = datetime.datetime.now(est).strftime('%Y-%m-%d')
        
        # 📊 [V24] 만약 ticker_state가 인자로 넘어오지 않았다면 live_status에서 파싱 시도
        # 하지만 스케줄러(main.py)에서 명시적으로 넘겨주는 것이 가장 정확함
        if ticker_state is None:
            live_status = self._load_json("LIVE_STATUS", {})
            ticker_state = {}
            if "tickers" in live_status:
                for t, info in live_status["tickers"].items():
                    ticker_state[t] = {
                        "qty": info.get("qty", 0),
                        "avg_price": info.get("avg_price", 0),
                        "current_price": info.get("current_price", 0),
                        "prev_close": info.get("prev_close", 0)
                    }
        
        # 성과 분석용 개별 종목 평가금 (레거시 UI 호환성 유지)
        ticker_eval = {}
        for t, info in ticker_state.items():
            eval_val = info.get("qty", 0) * info.get("current_price", 0)
            if eval_val > 0: ticker_eval[t] = round(eval_val, 2)
        
        if holdings_value <= 0 and ticker_eval:
            holdings_value = sum(ticker_eval.values())

        new_snap = {
            "date": today,
            "cash": round(total_cash, 2),
            "holdings": round(holdings_value, 2),
            "total": round(total_cash + holdings_value, 2),
            "ticker_eval": ticker_eval,
            "ticker_state": ticker_state # 🔥 [V24] 주말 복구용 핵심 데이터
        }
        
        idx = next((i for i, s in enumerate(snapshots) if s['date'] == today), -1)
        if idx >= 0:
            snapshots[idx] = new_snap
        else:
            snapshots.append(new_snap)
            
        self._save_json("SNAPSHOTS", snapshots)
        return new_snap

    def get_latest_ticker_state(self):
        """🚀 [V24 Stable] 휴면 시 대시보드 복구용 최신 박제 데이터 반환"""
        snapshots = self._load_json("SNAPSHOTS", [])
        if not snapshots:
            return {}
        # 가장 최근 날짜의 스냅샷에서 ticker_state 추출
        last_snap = snapshots[-1]
        return last_snap.get("ticker_state", {})


    # 🕒 [V23.1] 모의투자 LOC 에뮬레이션 전용 스테이징 시스템
    def stage_mock_loc_order(self, ticker, order):
        """🚀 [V23.1] 모의투자에서 지원하지 않는 LOC 주문을 장 마감 전 실행용으로 예약"""
        path = self._get_file_path("MOCK_LOC_STAGING")
        staged = self._load_json_from_path(path, [])
        
        # 중복 제거 (티커와 사이드 기준)
        staged = [o for o in staged if not (o['ticker'] == ticker and o['side'] == order['side'])]
        
        order['ticker'] = ticker
        staged.append(order)
        self._save_json_to_path(path, staged)
        logging.info(f"🕒 [MOCK-LOC] {ticker} {order['side']} 주문이 장 마감 전 실행을 위해 예약되었습니다.")

    def get_staged_mock_loc_orders(self):
        path = self._get_file_path("MOCK_LOC_STAGING")
        return self._load_json_from_path(path, [])

    def clear_staged_mock_loc_orders(self):
        path = self._get_file_path("MOCK_LOC_STAGING")
        self._save_json_to_path(path, [])
        logging.info("🧹 [MOCK-LOC] 모든 예약 주문이 초기화되었습니다.")


    def _load_json_from_path(self, path, default):
        if not path or not os.path.exists(path): return default
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except: return default

    def _save_json_to_path(self, path, data):
        if not path: return
        dir_name = os.path.dirname(path)
        if not os.path.exists(dir_name): os.makedirs(dir_name, exist_ok=True)
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except: pass

    # 📋 [V33] Unified Logging & Search System
    _last_logs = {} # Throttling cache: {(category, task, msg): timestamp}

    def log_event(self, category, task, status, message, details=None, is_alert=None):
        """
        🚀 [V33 Unified] 전체 시스템 통합 로그 기록기
        카테고리: TRADE(매매), SCHEDULE(일정), SYSTEM(시스템), SNAPSHOT(분석)
        상태: SUCCESS, ERROR, WARNING, INFO, STATUS, WAIT
        """
        import json
        est = pytz.timezone('US/Eastern')
        now_ny = datetime.datetime.now(est)
        
        # 1. 🛡️ 중복 로그 Throttling (동일 메시지 15분 내 반복 시 무시 - TRADE/ERROR 제외)
        if category.upper() not in ['TRADE', 'ERROR']:
            cache_key = (category.upper(), task.upper() if task else "SYSTEM", message)
            last_time = self._last_logs.get(cache_key)
            if last_time and (now_ny.timestamp() - last_time < 900): # 15분(900초)
                return
            self._last_logs[cache_key] = now_ny.timestamp()

        # 2. Master Event Log 기록 (event_log.json)
        events = self._load_json("EVENT_LOG", [])
        if len(events) > 2000: events = events[-2000:]
        
        entry = {
            "date": now_ny.strftime("%m/%d"),
            "time": now_ny.strftime("%H:%M:%S"),
            "category": category.upper(),
            "task": task.upper() if task else "SYSTEM",
            "status": status.upper(),
            "msg": message,
            "details": details
        }
        events.append(entry)
        self._save_json("EVENT_LOG", events)

        # 3. Notification Feed 기록 (유저 실시간 알림용)
        # 🚀 [V33.5] 실시간 피드 제약: 매매/스케줄/중요 오류만 기록 (일반 시스템 로그 제외)
        # 🚀 [V33.5] 수량 제한 제거 (사용자 수동 비우기 가능)
        if is_alert is None:
            is_alert = (category.upper() in ['TRADE', 'SCHEDULE', 'SYNC']) or (status.upper() in ['ERROR', 'WARNING', 'STATUS'])
        
        if is_alert:
            icons = {"SUCCESS": "✅", "ERROR": "🚨", "WARNING": "⚠️", "INFO": "📡", "WAIT": "⏳", "STATUS": "🔔"}
            notif_entry = {
                "date": entry["date"],
                "time": entry["time"],
                "task": entry["task"],
                "status": entry["status"],
                "msg": entry["msg"],
                "icon": icons.get(entry["status"], "📝")
            }
            
            notifs = self._load_json("NOTIFICATIONS", [])
            notifs.append(notif_entry)
            # 🚀 [V33.5] 무제한 저장
            self._save_json("NOTIFICATIONS", notifs)
    
    # [V33 Compatible] 기존 레거시 함수들 래핑
    def record_event(self, task_name, status, message, details=None):
        cat = "SYSTEM"
        uname = task_name.upper() if task_name else "SYSTEM"
        if "TRADE" in uname or "BUY" in uname or "SELL" in uname: cat = "TRADE"
        elif uname in ["SYNC", "RESET", "PRE", "REG", "AFTER", "IDLE"]: cat = "SCHEDULE"
        self.log_event(cat, task_name, status, message, details)

    def add_notification(self, level, message, phase="NONE", sync=False):
        cat = "SCHEDULE" if phase != "NONE" else "SYSTEM"
        if "TRADE" in phase.upper() or "완료" in message: cat = "TRADE"
        self.log_event(cat, phase, level, message, is_alert=True)

